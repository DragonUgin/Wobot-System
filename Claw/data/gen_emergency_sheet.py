# -*- coding: utf-8 -*-
"""生成《Claw 应急人工统计表》—— 断网/系统故障时，工作人员可打印后手写勾选的纸质台账。
所有静态数据（玩家/线索/露水/任务点/猎人）从现有表预填，只留状态栏手填。
输出：D:/58系统/Claw/data/应急人工统计表.xlsx
重跑即可重建。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "D:/58系统/Claw/data/"
OUT  = BASE + "应急人工统计表.xlsx"

# ---------- 读源数据 ----------
players = []   # (组号, 姓名)
wb = openpyxl.load_workbook(BASE + "玩家名单.xlsx")
ws = wb["玩家"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] == "玩家" and r[1] is not None and r[2]:
        players.append((int(r[1]), str(r[2])))

# 线索
clues = {}   # id -> (type, tp, hunter)
wb = openpyxl.load_workbook(BASE + "线索编号表.xlsx")
ws = wb["线索"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]:
        continue
    cid = int(r[1]); ctype = str(r[2] or ""); tp = r[3]; hunter = r[6]
    clues[cid] = (ctype, (int(tp) if tp not in (None, 0, "0") else 0), (str(hunter) if hunter else ""))

# 露水 + 猎人持有
dews = {}        # id -> hunter_name (from 露水编号表)
wb = openpyxl.load_workbook(BASE + "露水编号表.xlsx")
ws = wb["露水"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]:
        continue
    did = int(r[1]); hunter = r[2]
    dews[did] = (str(hunter) if hunter else "")

# 猎人持有露水（最终编号映射，来自导入总表 普通露水+归属猎人）
hunter_dews = {}   # hunter -> [dew_ids]
wb = openpyxl.load_workbook(BASE + "导入总表.xlsx")
ws = wb["导入总表"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] == "普通露水" and r[2]:
        did = int(r[1]); h = str(r[2])
        hunter_dews.setdefault(h, []).append(did)

# 任务点
tps = []   # (id, name, golden, shield, static)
wb = openpyxl.load_workbook(BASE + "导入总表.xlsx")
ws = wb["导入总表"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] == "任务点":
        tps.append((int(r[1]), str(r[2]), int(r[6] or 0), int(r[7] or 0), int(r[8] or 0)))

# 派生：每个任务点的线索目标数、露水目标数
def clue_target(tp):
    return sum(1 for c in clues.values() if c[1] == tp)
def dew_target(tp):
    return sum(1 for cid, c in clues.items() if c[0] == "真" and c[1] == tp)

# ---------- 样式 ----------
CB = "☐"   # 勾选框
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
HEAD_FILL  = PatternFill("solid", fgColor="D9E1F2")
HEAD_FONT  = Font(bold=True, size=10)
CELL_FONT  = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = CENTER; cell.border = BORDER

def box(title, ncols, widths, landscape=False, title_rows="1:2"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=f"《Claw 应急人工统计表》— {title}（断网可用 · 手写勾选）")
    t.fill = TITLE_FILL; t.font = TITLE_FONT; t.alignment = CENTER
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.print_title_rows = title_rows
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb, ws

# ================= 1. 玩家状态 =================
wb, ws = box("玩家状态", 8, [6, 12, 8, 8, 8, 16, 14, 18])
hdr = ["组号", "姓名", CB + "存活", CB + "淘汰", CB + "复活", "淘汰时间", "被谁淘汰", "备注"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=2, column=i, value=h)
style_header(ws, 2, 8)
players.sort(key=lambda x: (x[0], x[1]))
for ri, (g, name) in enumerate(players, 3):
    ws.cell(row=ri, column=1, value=g).alignment = CENTER
    ws.cell(row=ri, column=2, value=name).alignment = LEFT
    for c in (3, 4, 5):
        ws.cell(row=ri, column=c, value=CB).alignment = CENTER
    for c in (6, 7, 8):
        ws.cell(row=ri, column=c, value="").alignment = LEFT
    for c in range(1, 9):
        ws.cell(row=ri, column=c).font = CELL_FONT
        ws.cell(row=ri, column=c).border = BORDER
    ws.row_dimensions[ri].height = 16

# ================= 2. 任务点物资 =================
wb2, ws2 = box("任务点物资", 10, [8, 12, 9, 9, 8, 8, 8, 9, 12, 22], landscape=True)
hdr2 = ["任务点", "名称", "线索目标", "露水目标", "金露水", "护盾卡", "静止卡", CB + "已完成", "完成时间", "接收/送出记录"]
for i, h in enumerate(hdr2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header(ws2, 2, 10)
for ri, (tid, name, gold, sh, st) in enumerate(tps, 3):
    vals = [tid, name, clue_target(tid), dew_target(tid), gold, sh, st, CB, "", ""]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = LEFT if ci in (2, 9, 10) else CENTER
    ws2.row_dimensions[ri].height = 18

# ================= 3. 线索追踪 =================
wb3, ws3 = box("线索追踪", 7, [7, 8, 18, 9, 9, 9, 14])
hdr3 = ["编号", "类型", "关联任务点/猎人", CB + "未收集", CB + "已发现", CB + "已收集", "收集时间"]
for i, h in enumerate(hdr3, 1):
    ws3.cell(row=2, column=i, value=h)
style_header(ws3, 2, 7)
def rel_label(tp, hunter):
    parts = []
    if tp: parts.append(f"TP{tp}")
    if hunter: parts.append(f"猎人{hunter}")
    return " / ".join(parts) if parts else "无主"
for ri, cid in enumerate(sorted(clues), 3):
    ctype, tp, hunter = clues[cid]
    vals = [cid, ctype, rel_label(tp, hunter), CB, CB, CB, ""]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = LEFT if ci == 3 else CENTER
    ws3.row_dimensions[ri].height = 16

# ================= 4. 露水追踪 =================
wb4, ws4 = box("露水追踪", 6, [7, 20, 9, 9, 9, 14])
hdr4 = ["编号", "关联任务点/猎人", CB + "未收集", CB + "已发现", CB + "已收集", "收集时间"]
for i, h in enumerate(hdr4, 1):
    ws4.cell(row=2, column=i, value=h)
style_header(ws4, 2, 6)
for ri, did in enumerate(sorted(dews), 3):
    hunter = dews[did]
    tp = clues.get(did, ("", 0, ""))[1] if did in clues else 0
    vals = [did, rel_label(tp, hunter), CB, CB, CB, ""]
    for ci, v in enumerate(vals, 1):
        cell = ws4.cell(row=ri, column=ci, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = LEFT if ci == 2 else CENTER
    ws4.row_dimensions[ri].height = 16

# ================= 5. 猎人持有 =================
hunters = sorted(hunter_dews.items(), key=lambda kv: min(kv[1]))
wb5, ws5 = box("猎人持有", 4, [14, 18, 10, 24])
hdr5 = ["猎人", "持有露水编号", CB + "在册", "备注"]
for i, h in enumerate(hdr5, 1):
    ws5.cell(row=2, column=i, value=h)
style_header(ws5, 2, 4)
for ri, (h, dl) in enumerate(hunters, 3):
    vals = [h, "/".join(map(str, sorted(dl))), CB, ""]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(row=ri, column=ci, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = LEFT if ci in (1, 2, 4) else CENTER
    ws5.row_dimensions[ri].height = 16

# ================= 6. 操作记录 =================
wb6, ws6 = box("操作记录", 4, [16, 40, 12, 24])
hdr6 = ["时间", "事件（淘汰/复活/任务点完成/露水收集…）", "操作人", "备注"]
for i, h in enumerate(hdr6, 1):
    ws6.cell(row=2, column=i, value=h)
style_header(ws6, 2, 4)
for ri in range(3, 40):   # 空白可写区
    for ci in range(1, 5):
        ws6.cell(row=ri, column=ci, value="").border = BORDER
        ws6.cell(row=ri, column=ci).font = CELL_FONT
        ws6.cell(row=ri, column=ci).alignment = LEFT
    ws6.row_dimensions[ri].height = 16

# ================= 汇总到一个文件 =================
out = openpyxl.Workbook()
out.remove(out.active)
for w in (wb, wb2, wb3, wb4, wb5, wb6):
    s = w.active
    out.create_sheet(title=s.title)
    out_s = out[s.title]
    # 复制
    for row in s.iter_rows():
        for cell in row:
            out_s.cell(row=cell.row, column=cell.column, value=cell.value)
    out_s.row_dimensions[1].height = s.row_dimensions[1].height
    out_s.freeze_panes = s.freeze_panes
    out_s.print_title_rows = s.print_title_rows
    out_s.page_setup.orientation = s.page_setup.orientation
    out_s.page_setup.paperSize = s.page_setup.paperSize
    out_s.page_setup.fitToWidth = s.page_setup.fitToWidth
    out_s.page_setup.fitToHeight = s.page_setup.fitToHeight
    try:
        out_s.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    for col, dim in s.column_dimensions.items():
        out_s.column_dimensions[col].width = dim.width

out.save(OUT)
print("SAVED", OUT)
print("sheets:", out.sheetnames)
print("players:", len(players), "clues:", len(clues), "dews:", len(dews),
      "hunters:", len(hunter_dews), "taskpoints:", len(tps))
print("clue targets:", {t[0]: clue_target(t[0]) for t in tps})
print("dew  targets:", {t[0]: dew_target(t[0]) for t in tps})
