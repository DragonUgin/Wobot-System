# -*- coding: utf-8 -*-
"""整合 4 份已生成数据 + 任务点元数据 → 一份 导入总表.xlsx
格式对齐 utils/excel_importer.py（单 Sheet 一体式，第1行表头 A1=种类 会被跳过）。
"""
import sys, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8")

DATA = r"D:/58系统/Claw/data"
SRC_PLAYERS = DATA + "/玩家名单.xlsx"
SRC_CLUES   = DATA + "/线索编号表.xlsx"
SRC_DEWS    = DATA + "/露水编号表.xlsx"
DST         = DATA + "/导入总表.xlsx"

HDR_FILL = PatternFill("solid", fgColor="2F3A4A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(size=10, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

# ---- 任务点数据（用户确认：名称=任务点N，类型=solo，功能卡=无，静止卡5，护盾卡5）----
TP_DATA = [(1,3),(2,3),(3,0),(4,6),(5,0),(6,0)]  # (编号, 金露水数量)

# ---- 读取玩家 ----
def load_players():
    wb = openpyxl.load_workbook(SRC_PLAYERS, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or row[0] is None: continue
        if str(row[0]).strip() != "玩家": continue
        grp = row[1]; name = row[2]
        if grp is None or name is None: continue
        out.append((int(grp), str(name).strip()))
    wb.close()
    return out

# ---- 读取线索 ----
def load_clues():
    wb = openpyxl.load_workbook(SRC_CLUES, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or row[0] is None: continue
        if str(row[0]).strip() != "线索": continue
        cid = int(row[1]); ctype = str(row[2]).strip() if row[2] else "真"
        tp = int(row[3]) if row[3] is not None else 0
        out.append((cid, ctype, tp))
    wb.close()
    return out

# ---- 读取普通露水 ----
def load_dews():
    wb = openpyxl.load_workbook(SRC_DEWS, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or row[0] is None: continue
        if str(row[0]).strip() != "普通露水": continue
        did = int(row[1])
        hunter = str(row[2]).strip() if (len(row) > 2 and row[2]) else None
        out.append((did, hunter))
    wb.close()
    return out

# ---- 金露水（从 TP_DATA 派生：1-3→TP1, 4-6→TP2, 7-12→TP4）----
def gen_golden():
    out = []; gid = 1
    for tp, g in TP_DATA:
        for _ in range(g):
            out.append((gid, tp)); gid += 1
    return out

players = load_players()
clues   = load_clues()
dews    = load_dews()
goldens = gen_golden()
print("loaded: 玩家=%d 线索=%d 露水=%d 金露水=%d" % (len(players), len(clues), len(dews), len(goldens)))

# ---- 写总表 ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "导入总表"
header = ["种类","字段1","字段2","字段3","字段4","字段5","字段6","字段7","字段8"]
ws.append(header)
for c in range(1, 10):
    cell = ws.cell(row=1, column=c)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
ws.freeze_panes = "A2"

def pad(row9):  # 保证 9 列
    return row9 + [None]*(9-len(row9))

# 1. 任务点
for tp, g in TP_DATA:
    ws.append(pad(["任务点", tp, "任务点%d" % tp, "solo", "无", None, g, 5, 5]))
# 2. 玩家
for grp, name in players:
    ws.append(pad(["玩家", grp, name]))
# 3. 线索
for cid, ctype, tp in clues:
    ws.append(pad(["线索", cid, ctype, tp]))
# 4. 普通露水
for did, hunter in dews:
    ws.append(pad(["普通露水", did, hunter]))
# 5. 金露水
for gid, tp in goldens:
    ws.append(pad(["金露水", gid, tp]))

for col, w in zip("ABCDEFGHI", [10, 10, 14, 12, 12, 12, 14, 12, 12]):
    ws.column_dimensions[col].width = w

# ---- 说明 Sheet ----
ws2 = wb.create_sheet("说明")
ws2.cell(row=1, column=1, value="导入总表说明").font = Font(bold=True, size=13)
total = 6 + len(players) + len(clues) + len(dews) + len(goldens)
notes = [
    "本表为单 Sheet 一体式，第1行为表头（A1=种类，importer 自动跳过），第2行起为数据，共 %d 行。" % total,
    "字段映射（按种类，抄 excel_importer.py 文档）：",
    "  种类=任务点  : 字段1=编号 | 字段2=名称 | 字段3=类型(solo) | 字段4=功能卡 | 字段5=NPC_QQ(废弃) | 字段6=金露水数量 | 字段7=静止卡库存 | 字段8=护盾卡库存",
    "  种类=玩家    : 字段1=组号 | 字段2=姓名 | 字段3=备注",
    "  种类=线索    : 字段1=编号 | 字段2=类型(真/假) | 字段3=关联任务点 | 字段4=内容 | 字段5=藏于NPC | 字段6=归属猎人(可选)",
    "  种类=普通露水: 字段1=编号 | 字段2=归属猎人(可选,空=任务点池)",
    "  种类=金露水  : 字段1=编号 | 字段2=关联任务点",
    "本表组成：任务点6 + 玩家%d + 线索%d + 普通露水%d + 金露水%d = %d 行。" % (
        len(players), len(clues), len(dews), len(goldens), total),
    "导入流程：1) 设置→数据管理→清空所有数据；2) 导入Excel 选本文件；3) 核对各列表数量。",
    "线索全为「真」，编号41-53无对应任务点池露水(40张)，收集时不揭示，无副作用；如需假线索导入后网页改。",
    "猎人未单独导入（去名单化）；猎人池露水501-514已携带归属猎人名。",
]
for i, n in enumerate(notes, start=2):
    c = ws2.cell(row=i, column=1, value=n)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 110

wb.save(DST)
print("导入总表.xlsx -> %d 行 (任务点6 + 玩家%d + 线索%d + 露水%d + 金露水%d)" % (
    total, len(players), len(clues), len(dews), len(goldens)))
print("DONE")
