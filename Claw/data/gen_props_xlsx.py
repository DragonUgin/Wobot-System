# -*- coding: utf-8 -*-
"""把 金露水 / 护盾卡 / 静止卡 合做一张表（一个 Excel 文件）。
- 汇总表：按任务点列出 金露水数量+编号区间、护盾卡数量、静止卡数量
- 金露水编号明细：12 张金露水逐条编号（可导入 种类=金露水）
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HDR_FILL = PatternFill("solid", fgColor="2F3A4A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="2F3A4A")
SUB_FONT = Font(bold=True, size=11, color="2F3A4A")
NOTE_FONT = Font(size=10, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

# 任务点道具数据（来自用户名单）
# (任务点, 金露水数量, 护盾卡数量, 静止卡数量)
TP_DATA = [
    (1, 3, 5, 5),
    (2, 3, 5, 5),
    (3, 0, 5, 5),
    (4, 6, 5, 5),
    (5, 0, 5, 5),
    (6, 0, 5, 5),
]

# 金露水编号：按任务点顺序连续编号
golden = []  # (编号, 关联任务点)
gid = 1
for tp, g, s, t in TP_DATA:
    for _ in range(g):
        golden.append((gid, tp))
        gid += 1
# 每个任务点的金露水编号区间
ranges = {}
for num, tp in golden:
    ranges.setdefault(tp, []).append(num)
range_str = {tp: (f"{min(v)}-{max(v)}" if len(v) > 1 else f"{v[0]}") for tp, v in ranges.items()}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "道具汇总"

# 标题
ws.merge_cells("A1:E1")
ws["A1"] = "金露水 / 护盾卡 / 静止卡 汇总表"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = CENTER

# 汇总表头
hdr = ["任务点", "金露水(张)", "金露水编号", "护盾卡(张)", "静止卡(张)"]
ws.append([])  # row2 empty for spacing? no, write header at row2
for c, h in enumerate(hdr, start=1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER

# 汇总数据
r = 3
for tp, g, s, t in TP_DATA:
    row = [tp, g, range_str.get(tp, "—"), s, t]
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = CENTER; cell.border = BORDER
    r += 1
# 合计行
ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
ws.cell(row=r, column=1).alignment = CENTER
ws.cell(row=r, column=2, value=sum(x[1] for x in TP_DATA)).font = Font(bold=True)
ws.cell(row=r, column=4, value=sum(x[2] for x in TP_DATA)).font = Font(bold=True)
ws.cell(row=r, column=5, value=sum(x[3] for x in TP_DATA)).font = Font(bold=True)
for c in (1, 2, 4, 5):
    ws.cell(row=r, column=c).alignment = CENTER
    ws.cell(row=r, column=c).border = BORDER

# 金露水编号明细（隔一行）
r += 2
ws.cell(row=r, column=1, value="金露水编号明细（共 %d 张，可导入 种类=金露水）" % len(golden)).font = SUB_FONT
r += 1
ghdr = ["编号", "关联任务点", "", "", ""]
for c, h in enumerate(["编号", "关联任务点"], start=1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
r += 1
for num, tp in golden:
    ws.cell(row=r, column=1, value=num).alignment = CENTER
    ws.cell(row=r, column=2, value=tp).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER
    r += 1

for col, w in zip("ABCDE", [10, 12, 14, 12, 12]):
    ws.column_dimensions[col].width = w

# 说明 sheet
ws2 = wb.create_sheet("说明")
ws2.cell(row=1, column=1, value="说明").font = Font(bold=True, size=13)
notes = [
    "本表合并 金露水 / 护盾卡 / 静止卡 三种道具，按任务点汇总。",
    "金露水共 %d 张：任务点1=3(编号1-3)、任务点2=3(编号4-6)、任务点4=6(编号7-12)。" % len(golden),
    "金露水为独立编号道具，收集/完成任务点后送出；导入用 种类=金露水，字段=编号|关联任务点（见本表明细）。",
    "护盾卡 / 静止卡 为任务点库存数量（非单独编号道具），对应任务点完成/接收指令里的 护盾卡Z 静止卡Z。",
    "任务点名单中护盾卡均为5、静止卡均为5；金露水按名单分别为 3/3/0/6/0/0。",
    "游戏时长「持续2小时」未纳入本表，属游戏配置项，可在后台另行设置。",
]
for i, n in enumerate(notes, start=2):
    c = ws2.cell(row=i, column=1, value=n)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 95

wb.save(r"D:\58系统\Claw\data\金露水护盾卡静止卡.xlsx")
print("金露水护盾卡静止卡.xlsx -> 汇总表 + 金露水明细(%d张) OK" % len(golden))
