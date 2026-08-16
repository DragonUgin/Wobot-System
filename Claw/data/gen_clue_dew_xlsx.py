# -*- coding: utf-8 -*-
"""按最新指令清单与后端模型，生成两份编号 Excel：
   - 线索编号表.xlsx  (种类=线索)
   - 露水编号表.xlsx  (种类=普通露水)
格式对齐 utils/excel_importer.py（单 Sheet 一体式，第1行为表头，第2行起为数据）。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HDR_FILL = PatternFill("solid", fgColor="2F3A4A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(size=10, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.freeze_panes = "A2"

def add_note_sheet(wb, title, lines):
    ws = wb.create_sheet("说明")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    for i, ln in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=ln)
        c.font = NOTE_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 90

# ============ 数据来源 ============
# 任务点线索数：TP1~5 各6，TP6 为9 -> 共 39
TP_CLUES = {1: 6, 2: 6, 3: 6, 4: 6, 5: 6, 6: 9}
OTHER_CLUES = 15               # 无主线索 40-54（编号到54合法）

# 露水：总54张（编号1-54，与线索同号），其中 14 张归属猎人。
# 猎人露水编号与线索编号同号，且部分与任务点真线索编号重叠（31/33/35/37=任务点6），
# 实现"一个露水同时关联任务点和猎人"。add_dew 已支持按编号合并猎人归属。
HUNTER_DEW = {
    31: "示例姓名", 33: "示例姓名", 35: "示例姓名", 37: "示例姓名",
    40: "示例姓名", 41: "示例姓名", 42: "示例姓名", 43: "示例姓名",
    49: "示例姓名", 50: "示例姓名", 51: "示例姓名", 52: "示例姓名",
    53: "示例姓名", 54: "示例姓名",
}

# ============ 1. 线索编号表 ============
wb_c = openpyxl.Workbook()
ws_c = wb_c.active
ws_c.title = "线索"
ws_c.append(["种类", "编号", "类型(真/假)", "关联任务点", "内容", "藏于NPC", "归属猎人(可选)"])
cid = 1
for tp, cnt in TP_CLUES.items():
    for _ in range(cnt):
        ws_c.append(["线索", cid, "真", tp, "", "", ""])
        cid += 1
for _ in range(OTHER_CLUES):
    ws_c.append(["线索", cid, "真", 0, "", "", ""])   # 其他线索：无主，可后续分配
    cid += 1
style_header(ws_c, 7)
for col, w in zip("ABCDEFG", [8, 8, 12, 12, 30, 12, 16]):
    ws_c.column_dimensions[col].width = w
add_note_sheet(wb_c, "线索编号说明", [
    f"共 {cid-1} 条线索：任务点线索 39 条（任务点1~5各6，任务点6为9）+ 无主线索 15 条（40-54）。",
    "编号规则：按任务点顺序连续编号 —— 任务点1:1-6，任务点2:7-12，任务点3:13-18，任务点4:19-24，任务点5:25-30，任务点6:31-39，无主线索:40-54（编号到54合法）。",
    "类型(真/假)：默认全填「真」。真线索收集后会揭示同编号的露水（见露水编号表 1-54）。请按实际游戏设计把部分改为「假」。",
    "关联任务点：无主线索填 0，可在后台网页分配给任务点，或留作猎人持有物（填归猎人）。",
    "内容 / 藏于NPC / 归属猎人：留空待填（藏于NPC 改名后纯文本名字，去名单化；归属猎人 填了即归该猎人）。",
    "导入：直接用本 Sheet（线索）通过网页「导入Excel」即可，系统按 种类=线索 解析。",
])

# ============ 2. 露水编号表 ============
wb_d = openpyxl.Workbook()
ws_d = wb_d.active
ws_d.title = "露水"
ws_d.append(["种类", "编号", "归属猎人(可选)"])
# 露水 1-54（与线索同号）。14 张填归属猎人（编号 31/33/35/37/40/41/42/43/49/50/51/52/53/54），
# 其余不填 -> hunter_name=NULL。与同号真线索配对；猎人露水若对应任务点真线索则"同时关联任务点和猎人"。
for d in range(1, 55):
    ws_d.append(["普通露水", d, HUNTER_DEW.get(d, "")])
style_header(ws_d, 3)
for col, w in zip("ABC", [10, 8, 18]):
    ws_d.column_dimensions[col].width = w
add_note_sheet(wb_d, "露水编号说明", [
    "共 54 张露水（编号1-54），与线索编号一一对应（同号）。其中 14 张归属猎人（编号 31/33/35/37/40/41/42/43/49/50/51/52/53/54）。",
    "露水编号与线索编号同号：收集到同编号真线索即揭示该露水。",
    "归属猎人的露水（如露水31=示例姓名）：其对应线索若归属任务点（如线索31属任务点6），则该露水「同时关联任务点和猎人」；若对应线索无主（如露水40=示例姓名，线索40无主），则仅由猎人持有。",
    "任务点送出 / 猎人静止卡送出露水功能已砍掉（系统不再录入送出的露水）；露水存量仅可由网页端编辑，或经任务点/猎人「接收/存入」指令入库。",
    "add_dew 已支持「按编号合并猎人归属」（ON CONFLICT DO UPDATE）：导入时若同号露水已存在且本次指定猎人，则把猎人归属写入已有记录。",
    "导入：直接用本 Sheet（普通露水）通过网页「导入Excel」即可。",
])

wb_c.save(r"D:\58系统\Claw\data\线索编号表.xlsx")
wb_d.save(r"D:\58系统\Claw\data\露水编号表.xlsx")
print("线索编号表.xlsx ->", cid - 1, "条线索")
print("露水编号表.xlsx ->", 54, "张露水（其中", len(HUNTER_DEW), "张归属猎人）")
print("DONE")
