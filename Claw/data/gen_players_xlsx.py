# -*- coding: utf-8 -*-
# SPDX-License-Identifier: GPL-3.0-or-later
"""从 营员名单(1).xlsx 生成符合 Claw 导入格式的 玩家名单.xlsx
   导入格式：种类=玩家 | 组号 | 姓名 | 备注（第1行表头，第2行起数据）
"""
import sys, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"D:/Program Files (x86)/weixinDownload/xwechat_files/wxid_1370353703312_ac1b/msg/file/2026-08/营员名单(1).xlsx"
DST = r"D:/58系统/Claw/data/玩家名单.xlsx"

HDR_FILL = PatternFill("solid", fgColor="2F3A4A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(size=10, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

# 读取营员名单
src = openpyxl.load_workbook(SRC, read_only=True)
ws_src = src["Sheet1"]
players = []  # (组号, 姓名)
for i, row in enumerate(ws_src.iter_rows(values_only=True)):
    if i == 0:
        continue  # 表头
    if not row or row[0] is None:
        continue
    grp = row[1]
    name = row[2]
    if grp is None or name is None:
        continue
    try:
        grp = int(grp)
    except (TypeError, ValueError):
        continue
    name = str(name).strip()
    if not name:
        continue
    players.append((grp, name))
src.close()

# 组统计
from collections import Counter
grp_cnt = Counter(g for g, _ in players)

# 生成导入表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "玩家"
ws.append(["种类", "组号", "姓名", "备注"])
for grp, name in players:
    ws.append(["玩家", grp, name, ""])
# 表头样式
for c in range(1, 5):
    cell = ws.cell(row=1, column=c)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
ws.freeze_panes = "A2"
for col, w in zip("ABCD", [8, 8, 14, 16]):
    ws.column_dimensions[col].width = w

# 说明 sheet
ws2 = wb.create_sheet("说明")
ws2.cell(row=1, column=1, value="玩家名单说明").font = Font(bold=True, size=13)
notes = [
    "来源：营员名单(1).xlsx（Sheet1，126 名营员）。",
    "共 %d 名玩家，分布于 %d 个组：%s。" % (
        len(players), len(grp_cnt),
        "、".join("第%d组%d人" % (g, grp_cnt[g]) for g in sorted(grp_cnt))),
    "导入：直接用本 Sheet（玩家）通过网页「导入Excel」即可，系统按 种类=玩家 解析（组号|姓名|备注）。",
    "备注列留空，可自行填写（如身份/特殊标记），不影响导入。",
]
for i, n in enumerate(notes, start=2):
    c = ws2.cell(row=i, column=1, value=n)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 90

wb.save(DST)
print("玩家名单.xlsx -> %d 名玩家, %d 个组" % (len(players), len(grp_cnt)))
print("各组人数:", dict(sorted(grp_cnt.items())))
