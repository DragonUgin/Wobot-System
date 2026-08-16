"""生成测试用导入数据 sample_game_data.xlsx

对应 utils/excel_importer.py 的单 Sheet 一体式格式。
第 1 行必须是表头（导入时跳过），数据从第 2 行开始。

列定义：
  种类 | 字段1 | 字段2 | 字段3 | 字段4 | 字段5 | 字段6 | 字段7 | 字段8

  玩家     : 组号 | 姓名 | 备注
  NPC      : QQ号 | 名字 | 身份(task_npc/hunter/mobile)
  任务点   : 编号 | 名称 | 类型 | 功能卡 | [NPC_QQ已废弃·线下约定] | 金露水数 | 静止卡库存 | 护盾卡库存
  普通露水 : 编号
  金露水   : 编号 | 关联任务点
  线索     : 编号 | 类型(真/假) | 关联任务点 | 内容 | 藏于NPC姓名
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).resolve().parent / "sample_game_data.xlsx"

HEADER = ["种类", "字段1", "字段2", "字段3", "字段4", "字段5", "字段6", "字段7", "字段8"]

ROWS = [
    # ---- 玩家：组号 | 姓名 | 备注 ----
    ["玩家", 1, "示例姓名", "1组"],
    ["玩家", 1, "示例姓名", "1组"],
    ["玩家", 2, "示例姓名", "2组"],
    ["玩家", 2, "示例姓名", "2组"],

    # ---- NPC：QQ号 | 名字 | 身份 ----
    ["NPC", 0, "示例姓名", "hunter"],
    ["NPC", 0, "示例姓名", "hunter"],
    ["NPC", 0, "示例姓名", "task_npc"],
    ["NPC", 0, "示例姓名", "mobile"],

    # ---- 任务点：编号|名称|类型|功能卡|NPC_QQ|金露水|静止卡库存|护盾卡库存 ----
    ["任务点", 1, "示例姓名任务点", "solo", "护盾卡", "", 1, 1, 2],  # NPC_QQ列已废弃（线下口头约定）

    # ---- 普通露水：编号（每张 2 滴，仅真线索有对应露水）----
    ["普通露水", 1],
    ["普通露水", 3],
    ["普通露水", 4],

    # ---- 金露水：编号 | 关联任务点 ----
    ["金露水", 1, 1],

    # ---- 线索：编号|类型|关联任务点|内容|藏于NPC姓名 ----
    ["线索", 1, "真", 1, "真线索A：目标出现在图书馆三楼", "示例姓名"],
    ["线索", 2, "假", 1, "假线索B：目标已离开校区", "示例姓名"],
    ["线索", 3, "真", 1, "真线索C：目标携带蓝色背包", "示例姓名"],
    ["线索", 4, "真", 1, "真线索D：目标将在东门集合", "示例姓名"],
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入数据"

    ws.append(HEADER)
    head_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for r in ROWS:
        ws.append(r + [None] * (len(HEADER) - len(r)))

    for col, width in zip("ABCDEFGHI", (10, 14, 16, 12, 12, 26, 12, 12, 12)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"已生成 {OUT}")
    print(f"数据行 {len(ROWS)} 行：4玩家 / 4NPC / 1任务点 / 3普通露水 / 1金露水 / 4线索")


if __name__ == "__main__":
    main()
