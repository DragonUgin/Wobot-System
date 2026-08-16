"""Excel 导入 — 单 Sheet 一体式

重要：**第 1 行必须是表头行**（内容任意，会被跳过），数据从第 2 行开始。

格式（一行一个条目，种类列决定解析方式）：

种类 | 字段1 | 字段2 | 字段3 | 字段4 | 字段5 | 字段6 | 字段7 | 字段8

种类=玩家    : 组号 | 姓名 | 备注
种类=NPC     : QQ号 | 名字 | 身份(task_npc/hunter/mobile)
种类=任务点  : 编号 | 名称 | 类型(team_vs/team_coop/solo) | 功能卡 | NPC_QQ(多个用逗号分隔) | 金露水数量 | 静止卡库存 | 护盾卡库存
种类=普通露水: 编号 | 归属猎人(可选,填了则新建一条只属该猎人的露水,不填则入任务点池)
种类=金露水  : 编号 | 关联任务点
种类=线索    : 编号 | 类型(真/假) | 关联任务点 | 内容 | 藏于NPC_QQ | 归属猎人(可选,填了则归该猎人,不填且任务点为0则无主)
"""
import openpyxl

KINDS = {'玩家', 'NPC', '任务点', '普通露水', '金露水', '线索'}

async def import_all_from_excel(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active if wb.sheetnames else None

    result = {"玩家":0,"NPC":0,"任务点":0,"普通露水":0,"金露水":0,"线索":0,"errors":[]}

    if not ws:
        wb.close()
        return result

    # 容错：首行若已经是数据行（A列是合法种类），说明用户忘了写表头 —— 从第1行开始读，
    # 避免静默丢掉第一条记录。否则按约定跳过表头。
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    first_is_data = bool(first and first[0] and str(first[0]).strip() in KINDS)
    start_row = 1 if first_is_data else 2

    for i, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[0]: continue
        kind = str(row[0]).strip()
        if kind not in KINDS:
            result['errors'].append(f"第{i}行: 未知种类「{kind}」，已跳过")
            continue
        row_len = len(row)
        try:
            if kind == '玩家':
                if row_len < 3: result['errors'].append(f"第{i}行(玩家): 缺少组号或姓名"); continue
                g = int(row[1]); n = str(row[2]).strip()
                if n:
                    from utils.db import init_players
                    await init_players([(g, n)])
                    result['玩家'] += 1
            elif kind == 'NPC':
                if row_len < 2: result['errors'].append(f"第{i}行(NPC): 缺少QQ号"); continue
                qq = int(row[1]); name = str(row[2]).strip() if row_len > 2 else ""
                role = str(row[3]).strip() if row_len > 3 and row[3] else "hunter"
                from utils.db import add_npc
                await add_npc(qq, name, role)
                result['NPC'] += 1
            elif kind == '任务点':
                if row_len < 3: result['errors'].append(f"第{i}行(任务点): 缺少编号或名称"); continue
                tid = int(row[1]); name = str(row[2]).strip() if row_len > 2 else ""
                tp_type = str(row[3]).strip() if row_len > 3 and row[3] else "solo"
                card = str(row[4]).strip() if row_len > 4 and row[4] else ""
                # NPC_QQ 列(row[5])已废弃：任务点NPC名单砍掉，改线下口头约定，此列有数据也忽略
                golden = int(row[6]) if row_len > 6 and row[6] else 0
                static_cnt = int(row[7]) if row_len > 7 and row[7] else 0
                shield_cnt = int(row[8]) if row_len > 8 and row[8] else 0
                from utils.db import add_task_point
                await add_task_point(tid, name, tp_type, card, golden, static_cnt, shield_cnt)
                result['任务点'] += 1
            elif kind == '普通露水':
                if row_len < 2: result['errors'].append(f"第{i}行(普通露水): 缺少编号"); continue
                hunter = str(row[2]).strip() if row_len > 2 and row[2] else None
                from utils.db import add_dew
                await add_dew(int(row[1]), hunter)
                result['普通露水'] += 1
            elif kind == '金露水':
                if row_len < 2: result['errors'].append(f"第{i}行(金露水): 缺少编号"); continue
                gid = int(row[1]); tp_id = int(row[2]) if row_len > 2 and row[2] else 0
                from utils.db import add_golden_dew
                await add_golden_dew(gid, tp_id)
                result['金露水'] += 1
            elif kind == '线索':
                if row_len < 2: result['errors'].append(f"第{i}行(线索): 缺少编号"); continue
                cid = int(row[1]); ctype = str(row[2]).strip() if row_len > 2 and row[2] else "真"
                content = str(row[4]).strip() if row_len > 4 and row[4] else ""
                hidden_name = str(row[5]).strip() if row_len > 5 and row[5] else None
                hunter_name = str(row[6]).strip() if row_len > 6 and row[6] else None
                # 归属优先级：猎人名 > 关联任务点；都空 → 无主线索
                tp_id = 0 if hunter_name else (int(row[3]) if row_len > 3 and row[3] else 0)
                from utils.db import add_clue
                await add_clue(cid, ctype, tp_id, content, hidden_name, hunter_name)
                result['线索'] += 1
        except Exception as e:
            result['errors'].append(f"第{i}行({kind}): {e}")

    wb.close()
    return result
